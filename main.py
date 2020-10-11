# -*- coding: utf-8 -*-
"""
Created on Sun Oct  4 16:03:13 2020

@author: Q20712
"""

from openpyxl import load_workbook
import json
import datetime

import funcs

"""
 Settings
"""

#InputData
json_open = open('inputdata.json', 'r')
json_load = json.load(json_open)

thisMonth = True #Trueの場合、期間は今月1~最終日になる
startDate = [2020,9,1] #↑をFalseにしたら反映される
endDate = [2020,10,31]
user_name = json_load["user_name"]
user_fullname = json_load["user_fullname"]
excel_path = json_load["excel_path"]

json_open.close()

#ConfigData
OUTLOOK_FORMAT = '%m/%d/%Y %H:%M'

"""
 Processing
"""

## Outlookから必要な予定情報を取得して、配列に格納する ##
# カレンダーイベントの取得 #
if thisMonth:
    startDate, endDate = funcs.getStartDateAndEndDateInThisMonth()
else:
    startDate = datetime.date(startDate[0],startDate[1],startDate[2])
    endDate = datetime.date(endDate[0],endDate[1],endDate[2])
calendarItems = funcs.getCalendarItemsFromTo(startDate,endDate)

# ★付き予定の抽出 #
starItems = funcs.extractStarItems(calendarItems,user_fullname)

# ★付き予定をExcel入力用データ配列への変換 #
plan_data = funcs.excelInputData(starItems)

# startDateからendDateの配列作成 #
excel_input_data = []
nowDate = startDate
while nowDate != endDate:
    val = '赤坂'
    if nowDate.strftime('%A')=='Saturday' or nowDate.strftime('%A')=='Sunday':
        val = ''
    datam = {'date':nowDate,'value':val}
    excel_input_data.append(datam)
    nowDate = nowDate.replace(day=nowDate.day+1)
val = '赤坂'
if nowDate.strftime('%A')=='Saturday' or nowDate.strftime('%A')=='Sunday':
    val = ''
datam = {'date':endDate,'value':val}
excel_input_data.append(datam)

# excel_input_dataの修正 #(この操作で、excel_input_dataに全ての入力すべきデータが入る)
for i in excel_input_data:
    for j in plan_data:
        if i['date']==j['date']:
            i['value']=j['value']
            break


## 配列情報を元にExcelへ書き込みを行う ##
# Excelシートのオープン #
if startDate.month < 10:
    sheet_name_s = str(startDate.year)+'0'+str(startDate.month)
else:
    sheet_name_s = str(startDate.year)+str(startDate.month)
wb = load_workbook(excel_path)
all_sheets = wb.sheetnames
sheet_name = [s for s in all_sheets if s.startswith(sheet_name_s)][0] #startDateの示す月のワークシートが入ってる
ws = wb.get_sheet_by_name(sheet_name)

# user_nameの行を探す #
user_row = 1
user_name_column = 2
while ws.cell(row=user_row,column=user_name_column).value != user_name:
    user_row += 1

# 予定を書き込む #
now_column = 3
for plan in excel_input_data:
    ws.cell(row=user_row,column=now_column,value=plan['value'])
    now_column += 1

# Excelシートのクローズ #
wb.save(excel_path)
wb.close()

"""
# プリント用フィールド名
calcTableHeader = ['Title', 'Organizer', 'Start', 'Duration(Minutes)']
calcTableBody = []

#暫定表示用処理（あとで消そうね）
for appointmentItem in starItems:
    row = []
    row.append(appointmentItem.Subject)
    row.append(appointmentItem.Organizer)
    row.append(appointmentItem.Start.Format(OUTLOOK_FORMAT))
    row.append(appointmentItem.Duration)
    calcTableBody.append(row)

print(tabulate(calcTableBody, headers=calcTableHeader))
"""